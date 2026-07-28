"""
Cash Income & Deposit Extractor
استخراج بيانات الكاش والإيداع من ملفات PDF وتصديرها إلى Excel
"""

import os
import re
import glob
import pdfplumber
import pandas as pd
from pathlib import Path
from datetime import datetime


# ============================================================
# إعدادات المسارات - عدّلها حسب موقع ملفاتك
# ============================================================
BASE_DIR = r"D:\ملفات السنة\ملف 2026\Month_6\Daily-income"
OUTPUT_DIR = r"D:\ملفات السنة\ملف 2026\Month_6"
OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, "CASH-Income-And-Deposit-6-New.xlsx")

# أرقام الأيام المراد معالجتها
DAY_NUMBERS = list(range(1, 32))  # من 1 إلى 31


# ============================================================
# دوال استخراج البيانات
# ============================================================

def extract_text_from_pdf(pdf_path):
    """استخراج النص الكامل من ملف PDF"""
    full_text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                # النص من اليمين لليسار (نظراً لأن الملف عربي)
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
        return full_text
    except Exception as e:
        print(f"❌ خطأ في قراءة {pdf_path}: {e}")
        return ""


def reverse_arabic_text(text):
    """عكس النص العربي وإصلاح ترتيب الحروف فيه ليُقرأ بشكل صحيح"""
    lines = text.split('\n')
    reversed_lines = []
    
    def rev_arabic(match):
        return match.group(0)[::-1]
        
    for line in lines:
        words = line.split()
        reversed_words = ' '.join(reversed(words))
        # عكس الحروف للكلمات العربية بعد عكس الكلمات
        fixed_line = re.sub(r'[\u0600-\u06FF]+', rev_arabic, reversed_words)
        reversed_lines.append(fixed_line)
        
    return '\n'.join(reversed_lines)


def extract_daily_cash_data(pdf_path, day_number):
    """استخراج بيانات الكاش والمصروفات ليوم معين"""
    
    text = extract_text_from_pdf(pdf_path)
    if not text:
        return None
    
    # عكس النص العربي
    text = reverse_arabic_text(text)
    
    data = {
        'day': day_number,
        'date': None,
        'invoices': [],
        'total_sales_cash': 0.0,
        'total_deposits': 0.0,
        'cash_remaining': 0.0,
        'cash_transfer_out': 0.0,
        'bank_transfer_out': 0.0,
        'expenses': []
    }
    
    # استخراج التاريخ
    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', text)
    if date_match:
        data['date'] = date_match.group(1)
        
    # استخراج الفواتير والمصروفات عبر الجداول لضمان دقة الكاش
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    is_expense_section = False
                    cash_col_idx = 4
                    desc_col_start = 5
                    
                    for row in table:
                        if not row:
                            continue
                            
                        # تحويل كافة الخلايا إلى نصوص لتسهيل التعامل
                        row_strs = [str(x).strip() if x else '' for x in row]
                        joined_row = ''.join(row_strs)
                        
                        # تفعيل وضع قراءة المصروفات إذا وجدنا كلمة المصروفات في الترويسة
                        if 'فيراصملا' in joined_row and 'دنب' in joined_row:
                            is_expense_section = True
                            continue
                            
                        # 1. قراءة الإيرادات (قبل قسم المصروفات، وبناءً على ترتيب أعمدة الإيرادات)
                        if not is_expense_section and len(row) >= 12:
                            num_str = row_strs[11]
                            if num_str.isdigit():
                                cash_val_str = row_strs[1].replace(',', '')  # عمود الكاش في الإيرادات هو 1
                                if cash_val_str and cash_val_str.lower() != 'none':
                                    try:
                                        amount = float(cash_val_str)
                                        if amount > 0:
                                            name_str = row_strs[10]
                                            reversed_name = reverse_arabic_text(name_str).replace('\n', ' ')
                                            
                                            invoice_num = ""
                                            customer_name = reversed_name
                                            match_num = re.search(r'(\d{3,})', reversed_name)
                                            if match_num:
                                                invoice_num = match_num.group(1)
                                                customer_name = reversed_name.replace(invoice_num, '').strip()
                                            
                                            data['invoices'].append({
                                                'invoice_number': invoice_num,
                                                'customer_name': customer_name,
                                                'amount1': amount,
                                                'amount2': amount
                                            })
                                    except ValueError:
                                        pass
                                        
                        # 2. قراءة المصروفات الكاش فقط (داخل قسم المصروفات)
                        if is_expense_section and len(row) > cash_col_idx:
                            last_val = row_strs[-1]
                            if last_val.isdigit():
                                cash_val_str = row_strs[cash_col_idx].replace(',', '')  # عمود الكاش المخصص للمصروفات
                                if cash_val_str and cash_val_str.lower() != 'none':
                                    try:
                                        amount = float(cash_val_str)
                                        if amount > 0:
                                            # تجميع نص المصروف وعكس اللغة العربية لجعله مقروءاً
                                            desc_parts = []
                                            for cell in row_strs[desc_col_start:-1]:
                                                if cell:
                                                    desc_parts.append(reverse_arabic_text(cell).replace('\n', ' '))
                                            description = ' '.join(desc_parts).strip()
                                            if not description:
                                                description = "مصروف نقدي عام"
                                                
                                            data['expenses'].append({
                                                'description': description,
                                                'amount': amount
                                            })
                                    except ValueError:
                                        pass
    except Exception as e:
        print(f"❌ خطأ في استخراج الجداول من ملف الـ PDF: {e}")

    # استخراج الإجماليات (باستخدام أنماط أكثر مرونة)
    total_sales_match = re.search(r'(?:ايرادات الكاش|إجمالي المبيعات|المبيعات).*?\s+(\d+\.?\d*)', text)
    if total_sales_match:
        data['total_sales_cash'] = float(total_sales_match.group(1))
    
    total_deposits_match = re.search(r'(?:مصروفات الكاش|إجمالي المصروفات|المصروفات).*?\s+(\d+\.?\d*)', text)
    if total_deposits_match:
        data['total_deposits'] = float(total_deposits_match.group(1))
    
    cash_remaining_match = re.search(r'(?:المتبقي كاش|الكاش المتبقي|المتبقي|نقدية).*?\s+(-?\d+\.?\d*)', text)
    if cash_remaining_match:
        data['cash_remaining'] = float(cash_remaining_match.group(1))
    
    return data


def find_pdf_file(day_number):
    """البحث عن ملف PDF في مجلد اليوم"""
    day_folder = os.path.join(BASE_DIR, str(day_number))
    if not os.path.exists(day_folder):
        return None
    
    # البحث عن أي ملف PDF في المجلد
    pdf_files = glob.glob(os.path.join(day_folder, "*.pdf"))
    if pdf_files:
        return pdf_files[0]
    return None


def process_all_days():
    """معالجة جميع الأيام وتجميع البيانات"""
    
    all_data = []
    summary = {
        'total_sales_cash': 0.0,
        'total_deposits': 0.0,
        'total_cash_remaining': 0.0,
    }
    
    print("=" * 60)
    print("🚀 بدء معالجة ملفات الكاش الشهرية")
    print("=" * 60)
    
    for day_num in DAY_NUMBERS:
        pdf_path = find_pdf_file(day_num)
        
        if not pdf_path:
            print(f"⚠️  اليوم {day_num}: لم يتم العثور على ملف")
            continue
        
        print(f"📄 اليوم {day_num}: {os.path.basename(pdf_path)}")
        
        data = extract_daily_cash_data(pdf_path, day_num)
        if data:
            all_data.append(data)
            summary['total_sales_cash'] += data['total_sales_cash']
            summary['total_deposits'] += data['total_deposits']
            summary['total_cash_remaining'] += data['cash_remaining']
    
    return all_data, summary


def create_excel_report(all_data, summary):
    """إنشاء ملف Excel موحد بنفس التنسيق المطلوب في الصورة"""
    
    output_path = OUTPUT_EXCEL
    
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "التقرير"
    # تفعيل اتجاه الورقة من اليمين لليسار
    ws.sheet_view.rightToLeft = True

    # تعريف التنسيقات
    header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    gray_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # كتابة العناوين (من اليمين لليسار)
    headers = ['التاريخ', 'رقم الفاتورة', 'استلام مبلغ اجور', 'صرف', 'بيان الصرف', 'ايداعات']
    ws.append(headers)
    
    for col_idx in range(1, 7):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border
        
    # تحديد عرض الأعمدة
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15

    total_income = 0
    total_expense = 0

    current_row = 2
    for day_idx, day_data in enumerate(all_data):
        date_str = day_data['date'] if day_data['date'] else f"اليوم {day_data['day']}"
        invoices = day_data['invoices']
        expenses = day_data['expenses']
        
        max_len = max(len(invoices), len(expenses))
        if max_len == 0:
            continue
            
        for i in range(max_len):
            row_vals = [""] * 6
            
            # عمود A: التاريخ (يظهر مع كل فاتورة، أو في أول سطر لليوم)
            if i < len(invoices) or i == 0:
                row_vals[0] = date_str
            
            # الفواتير: رقم الفاتورة (B) والمبلغ (C)
            if i < len(invoices):
                row_vals[1] = invoices[i]['invoice_number']
                row_vals[2] = invoices[i]['amount1']
                total_income += invoices[i]['amount1']
                
            # المصروفات: المبلغ (D) والبيان (E)
            if i < len(expenses):
                row_vals[3] = expenses[i]['amount']
                row_vals[4] = expenses[i]['description']
                total_expense += expenses[i]['amount']
            
            for col_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.alignment = center_align
                cell.border = border
                
            current_row += 1
            
        # إضافة صف فاصل رمادي بين الأيام
        if day_idx < len(all_data) - 1:
            for col_idx in range(1, 7):
                cell = ws.cell(row=current_row, column=col_idx, value="")
                cell.fill = gray_fill
                cell.border = border
            current_row += 1

    # إضافة صف الإجمالي بأسفل الجدول مباشرة
    ws.cell(row=current_row, column=1, value="الاجمالي").font = bold_font
    ws.cell(row=current_row, column=3, value=total_income).font = bold_font
    ws.cell(row=current_row, column=4, value=total_expense).font = bold_font
    
    for col_idx in range(1, 7):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.alignment = center_align
        cell.border = border
    
    current_row += 2
    
    # مربع الملخص الملون (أسفل الجدول)
    # ملاحظة: يمكنك تغيير قيمة الإيداع النقدي لتسحب من البيانات إن وجدت
    cash_deposit = 0 
    cash_remaining = total_income - total_expense - cash_deposit
    
    summary_data = [
        ("اجمالي استلام", total_income, green_fill),
        ("اجمالي الصرف", total_expense, red_fill),
        ("ايداع النقدي", cash_deposit, yellow_fill),
        ("كاش المتبقي", cash_remaining, None),
    ]
    
    for desc, val, fill_color in summary_data:
        cell_desc = ws.cell(row=current_row, column=1, value=desc)
        cell_val = ws.cell(row=current_row, column=2, value=val)
        
        cell_desc.font = bold_font
        cell_val.font = bold_font
        
        cell_desc.alignment = center_align
        cell_val.alignment = center_align
        
        if fill_color:
            cell_desc.fill = fill_color
            cell_val.fill = fill_color
            
        current_row += 1

    wb.save(output_path)
    
    print(f"\n✅ تم إنشاء الملف بنجاح بالشكل المخصص: {output_path}")
    return output_path


# ============================================================
# الدالة الرئيسية
# ============================================================
def main():
    start_time = datetime.now()
    
    # معالجة جميع الأيام
    all_data, summary = process_all_days()
    
    # إنشاء ملف Excel
    if all_data:
        output_file = create_excel_report(all_data, summary)
        
        print("\n" + "=" * 60)
        print("📊 الملخص النهائي")
        print("=" * 60)
        print(f"عدد الأيام المعالجة : {len(all_data)}")
        print(f"إجمالي المبيعات     : {summary['total_sales_cash']:.2f} ريال")
        print(f"إجمالي المصروفات    : {summary['total_deposits']:.2f} ريال")
        print(f"الكاش المتبقي       : {summary['total_cash_remaining']:.2f} ريال")
        print(f"الملف المُنشأ        : {output_file}")
        print(f"الوقت المستغرق      : {datetime.now() - start_time}")
        print("=" * 60)
    else:
        print("❌ لم يتم العثور على أي بيانات")


if __name__ == "__main__":
    main()
